import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def fetch_stock_data(symbol):
    stock = yf.Ticker(symbol)
    df = stock.history(period="1y")  # Fetch 1 year of data
    if not df.empty:
        df = df[['Open', 'High', 'Low', 'Close', 'Volume']]
        df.index = df.index.tz_localize(None)  # Remove timezone information
        return df
    else:
        print(f"No data found for {symbol}")
        return None

def create_summary_sheet(writer, all_data):
    summary_data = []
    for symbol, data in all_data.items():
        latest_date = data.index[-1]
        latest_data = data.loc[latest_date]
        previous_data = data.loc[data.index[-2]]
        change = latest_data['Close'] - previous_data['Close']
        changeYüzde = 100 - ((latest_data['Close'] * 100) / previous_data['Close'])
        color = 'FF0000' if change < 0 else '00FF00'
        summary_data.append([symbol, latest_data['Open'], latest_data['High'], latest_data['Low'], latest_data['Close'], latest_data['Volume'], change, changeYüzde, color])

    summary_df = pd.DataFrame(summary_data, columns=['Symbol', 'Open', 'High', 'Low', 'Close', 'Volume', 'Change', 'Change %', 'Color'])

    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    return summary_df

def apply_colors(workbook, summary_df):
    ws = workbook['Summary']
    for idx, row in summary_df.iterrows():
        color = row['Color']
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=idx+2, column=col)  # +2 because openpyxl is 1-indexed and there's a header row
            cell.fill = fill

def main():
    # List of symbols for Borsa Istanbul stocks
    symbols = ['ASELS.IS', 'ALTNY.IS', 'AGROT.IS', 'KOPOL.IS', 'ALFAS.IS', 'ASTOR.IS', 'EUPWR.IS', 'CVKMD.IS', 'CWENE.IS', 'ENERY.IS', 'FZLGY.IS', 'GOKNR.IS', 'HEKTS.IS', 'KOCMT.IS', 'HRKET.IS', 'IZENR.IS', 'ODINE.IS', 'PLTUR.IS', 'KOTON.IS', 'LILAK.IS', 'LMKDC.IS', 'TNZTP.IS', 'REEDER.IS', 'SAHOL.IS', 'SDTTR.IS', 'SOKM.IS', 'VAKBN.IS']
    all_data = {}

    for symbol in symbols:
        print(f"Fetching data for {symbol}...")
        stock_data = fetch_stock_data(symbol)
        if stock_data is not None:
            all_data[symbol] = stock_data

    # Save the data to an Excel file
    if all_data:
        filename = f'stock_data_{datetime.now().strftime("%Y%m%d")}.xlsx'
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for symbol, data in all_data.items():
                data.to_excel(writer, sheet_name=symbol)
            
            # Create the summary sheet
            summary_df = create_summary_sheet(writer, all_data)

        # Apply colors to the summary sheet
        workbook = load_workbook(filename)
        apply_colors(workbook, summary_df)
        workbook.save(filename)
        print("Data successfully fetched and saved.")
    else:
        print("No data fetched, Excel file not created.")

if __name__ == "__main__":
    main()
